import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


def write_to_excel(fileData, fileName):
    dataColumns = [
        "Player", "Team",
        "15+ Points", "20+ Points", "25+ Points", "30+ Points",
        "4+ Reb", "6+ Reb", "8+ Reb", "10+ Reb", "12+ Reb",
        "4+ Assist", "6+ Assist", "8+ Assist", "10+ Assist", "12+ Assist",
        "2+ 3PM", "3+ 3PM", "4+ 3PM", "5+ 3PM"
    ]

    df = pd.DataFrame(fileData, columns=dataColumns)

    with (pd.ExcelWriter(fileName) as writer):
        df.to_excel(writer, sheet_name='espnPlayerData', index=False)

    format_excel(fileName)


def format_excel(fileName):
    # formattedfileName = './DataSheets/ESPN_PlayerData_' + d1 + '_FORMATTED.xlsx'
    workbook = load_workbook(fileName)
    sheet = workbook.active

    # Read the data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Create a new workbook for output
    output_workbook = Workbook()
    output_sheet = output_workbook.active

    # Write the data to the new sheet
    for row in data:
        output_sheet.append(row)

    # Determine the range for conditional formatting
    last_row = output_sheet.max_row

    if last_row < 2:
        print("Not enough rows for conditional formatting.")
    else:
        for column in output_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            output_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        formatting_range = f'C2:T{last_row}'
        color_scale = ColorScaleRule(start_type='percentile', start_value='0', start_color="F8696B",
                                     mid_type='percentile', mid_value='50', mid_color="FFEB84",
                                     end_type='percentile', end_value='100', end_color="63BE7B")

        output_sheet.conditional_formatting.add(formatting_range, color_scale)

    # Save the new workbook
    output_workbook.save(fileName)

