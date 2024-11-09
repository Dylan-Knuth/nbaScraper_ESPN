import requests
import datetime
import string
import time
import mysql.connector
import pyodbc

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from pyarrow import null

import PlayerTrendEmail
import NBAPlayer
import RequestTracker


def getTeamsPlayingToday():
    scrape_link = 'https://www.espn.com/nba/schedule'
    try:
        page = requests.get(scrape_link, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Accept": "*",
        })
        # Format the date as "Day, Month Date, Year"
        today = datetime.date.today().strftime("%A, %B %#d, %Y")
        teamsPlaying = []
        if page.status_code == 200:
            soup = BeautifulSoup(page.content, "html.parser")

            fullScheduleCard = soup.find("div", class_="Wrapper Card__Content overflow-visible")
            dateDivs = fullScheduleCard.find_all("div", class_="Table__Title")
            todaysGames = ""
            for date in dateDivs:
                dateText = date.get_text(strip=True).strip()
                if dateText == today:
                    todaysGames = date.find_next("table")
                    break

            if todaysGames:
                rows = todaysGames.find_all("tr", class_="Table__TR Table__TR--sm Table__even")
                for row in rows:
                    team_spans = row.find_all('span', class_='Table__Team')
                    for span in team_spans:
                        # Extract and append the text (team name) to the list
                        teamsPlaying.append(span.get_text(strip=True))

            print("Got teams playing")
            return teamsPlaying

    except Exception as e:
        print(e)


def espnScraper(nbaPlayer, tracker):
    playerLink = nbaPlayer.link
    scrape_link = 'https://www.espn.com/nba/player/gamelog/_/id/' + playerLink

    # try to access link
    try:
        page = requests.get(scrape_link, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
            "Accept": "*",
            "Connection": "close"
        })
        # add request to tracker
        tracker.add_request()

        howManyGames: int = 5

        # when page status is 200 scrape info
        if page.status_code == 200:
            soup = BeautifulSoup(page.content, "html.parser")

            playerName = string.capwords(str(playerLink.strip().split('/', 1)[1].replace('-', ' ')))

            # Get Team name and city
            teamDiv = soup.find("div", class_="PlayerHeader__Team")
            teamItem = teamDiv.find('a', class_="AnchorLink")
            if teamItem:
                team = teamItem.text
                num_spaces = team.count(" ")
                words = team.split()
                teamCity = ' '.join(team.split()[:num_spaces])
                teamName = words[-1]
            else:
                teamName = "Free Agent"
                teamCity = 'N/A'

            # Update Team info if required
            if teamName != nbaPlayer.teamName:
                nbaPlayer.teamName = teamName

            if teamCity != nbaPlayer.teamCity:
                nbaPlayer.teamCity = teamCity

            # Get injury status
            if soup.find("span", class_="TextStatus"):
                injuredStatus = soup.find("span", class_="TextStatus").get_text()
            else:
                injuredStatus = "Out Of League"
                nbaPlayer.status = injuredStatus
                return nbaPlayer

            # nbaPlayer = NBAPlayer.NBAPlayer(playerName, teamName, teamCity, '', injuredStatus)
            nbaPlayer.status = injuredStatus

            print(playerName + ": " + injuredStatus)

            seasonDivs = soup.find_all("div", class_='Table__Title')
            for seasonDiv in seasonDivs:
                if "Regular Season" in seasonDiv.text:
                    regularSeasonDiv = seasonDiv.previous_element
                    monthLogs = regularSeasonDiv.find_all("table", class_="Table Table--align-right")

                    for monthLog in monthLogs:
                        gameLogs = monthLog.tbody.find_all("tr", class_=["Table__TR Table__TR--sm Table__even",
                                                                         "filled Table__TR Table__TR--sm Table__even"])

                        for gameLog in gameLogs:
                            if gameLog:
                                gameCounter = len(nbaPlayer.games['points'])
                                if gameCounter < howManyGames or gameCounter < len(gameLogs):
                                    statBoxes = gameLog.find_all("td")
                                    nbaPlayer.add_game_stats(int(statBoxes[16].text.strip()),
                                                             int(statBoxes[10].text.strip()),
                                                             int(statBoxes[11].text.strip()),
                                                             int(statBoxes[6].text.strip().split('-', 1)[0]))
                            else:
                                nbaPlayer.add_game_stats("No point data",
                                                         "No rebound data",
                                                         "No assist data",
                                                         "No 3PT data")

            page.close()
            return nbaPlayer
        else:
            print(page.status_code)
            page.close()

    except Exception as e:
        print(e)


def write_to_excel(fileData, fileName):
    dataColumns = [
        "Player", "Team",
        "15+ Points", "20+ Points", "25+ Points", "30+ Points",
        "4+ Reb", "6+ Reb", "8+ Reb", "10+ Reb", "12+ Reb",
        "4+ Assist", "6+ Assist", "8+ Assist", "10+ Assist", "12+ Assist",
        "2+ 3PM", "3+ 3PM", "4+ 3PM", "5+ 3PM"
    ]

    df = pd.DataFrame(fileData, columns=dataColumns)

    with (pd.ExcelWriter(fileName) as writer):
        df.to_excel(writer, sheet_name='espnPlayerData', index=False)

    format_excel(fileName)


def format_excel(fileName):
    # formattedfileName = './DataSheets/ESPN_PlayerData_' + d1 + '_FORMATTED.xlsx'
    workbook = load_workbook(fileName)
    sheet = workbook.active

    # Read the data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Create a new workbook for output
    output_workbook = Workbook()
    output_sheet = output_workbook.active

    # Write the data to the new sheet
    for row in data:
        output_sheet.append(row)

    # Determine the range for conditional formatting
    last_row = output_sheet.max_row

    if last_row < 2:
        print("Not enough rows for conditional formatting.")
    else:
        for column in output_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            output_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        formatting_range = f'C2:T{last_row}'
        color_scale = ColorScaleRule(start_type='percentile', start_value='0', start_color="F8696B",
                                     mid_type='percentile', mid_value='50', mid_color="FFEB84",
                                     end_type='percentile', end_value='100', end_color="63BE7B")

        output_sheet.conditional_formatting.add(formatting_range, color_scale)

    # Save the new workbook
    output_workbook.save(fileName)


def get_NBA_Players():
    # Set up Selenium WebDriver (make sure to specify the path to your WebDriver)
    driver = webdriver.Chrome(executable_path='path/to/chromedriver')

    # Navigate to the ESPN NBA stats page
    url = "https://www.espn.com/nba/stats/player/_/table/general/sort/avgMinutes/dir/desc/"
    driver.get(url)

    # Click "Show More" until we have at least 200 rows
    while True:
        try:
            # Wait for the "Show More" button to be clickable and click it
            show_more_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Show More')]"))
            )
            show_more_button.click()

            # Wait for new rows to load
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//table/tbody/tr"))
            )

            # Check the number of rows in the table
            rows = driver.find_elements(By.XPATH, "//table/tbody/tr")
            if len(rows) >= 200:
                break

        except Exception as e:
            print(f"An error occurred: {e}")
            break

    # Scrape the data from the table
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    table = soup.find('table')
    data = []

    # Extract header information
    headers = [header.text for header in table.find_all('th')]
    data.append(headers)

    # Extract row data
    for row in table.find_all('tr')[1:]:
        cols = row.find_all('td')
        data.append([col.text.strip() for col in cols])

    # Convert to DataFrame and save to CSV or display
    df = pd.DataFrame(data[1:], columns=data[0])
    print(df)

    # Close the driver
    driver.quit()

    # scrape_link = 'https://www.espn.com/nba/stats/player/_/table/general/sort/avgMinutes/dir/desc/'
    # try:
    #     page = requests.get(scrape_link, headers={
    #         "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    #         "Accept": "*",
    #         "Connection": "close"
    #     })
    #
    #
    #     if page.status_code == 200:
    #         print("Yes")
    #
    # except Exception as e:
    #     print(e)

#
def get_NBAPlayers_DB(teamsPlaying):
    cnxn = pyodbc.connect(r'Driver=SQL Server;Server=.\SQLEXPRESS;Database=espnScraper;Trusted_Connection=yes;')
    cursor = cnxn.cursor()
    placeholders = tuple(teamsPlaying)

    cursor.execute("SELECT * FROM [espnScraper].[dbo].[NBAPlayers] WHERE teamCity in {}".format(placeholders))
    rows = cursor.fetchall()
    result_list = [list(row) for row in rows]
    cnxn.close()
    return result_list

def main():
    tracker = RequestTracker.RequestTracker()
    fileData = []
    playerData = []

    # get teams playing
    teamsPlaying = getTeamsPlayingToday()

    # get players from DB that are playing today
    ESPNPlayers = get_NBAPlayers_DB(teamsPlaying)

    d1 = datetime.datetime.now().strftime('%x').replace('/', '.')
    fileName = './DataSheets/ESPN_PlayerData_' + d1 + '.xlsx'

    print('Processing...........')
    if ESPNPlayers:
        for espnPlayer in ESPNPlayers:
            # parse players name from link
            playerData = []

            nbaPlayer = NBAPlayer.NBAPlayer(espnPlayer[0].strip(), espnPlayer[1].strip(),
                                            espnPlayer[4].strip(), espnPlayer[2].strip(), '?')
            nbaPlayer = espnScraper(nbaPlayer, tracker)
            player = nbaPlayer
            print(f"Requests in the last minute: {tracker.get_requests_per_minute()}")
            if player:
                if len(player.games['points']) > 0:
                    if player.status != 'Out':
                        allBenchmarks = player.get_all_benchmarks()
                        # player.print_benchmarks()
                        playerData.append(player.name)
                        playerData.append(player.teamCity + " " + player.teamName)

                        for stat, benchmarks in allBenchmarks.items():
                            for threshold, frequency in benchmarks.items():
                                playerData.append(frequency)

                        fileData.append(playerData)

                else:
                    print(f"{player.name} has no stats. {player.status}")
            else:
                print(f"Issue getting info for {espnPlayer}")
            time.sleep(5)

            # Add data to CSV file
    if fileData:
        write_to_excel(fileData, fileName)
        print("File saved to " + fileName)
        creds = PlayerTrendEmail.get_google_creds()
        PlayerTrendEmail.send_player_trends_email(creds, fileName)

    else:
        print("Issue with fileData. Cannot write to Excel")


if __name__ == "__main__":
    main()
